from __future__ import annotations
from xlrd import open_workbook
import os.path
import tempfile
import logging
from typing import TYPE_CHECKING, Iterator
import shutil
if TYPE_CHECKING:
    pass
import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from desiteRuleCreator.data import classes, constants
from desiteRuleCreator.Filehandling import open_file
from desiteRuleCreator.Windows import graphs_window
from desiteRuleCreator.Template import IFC_4_1

ident_pset_name = "Allgemeine Eigenschaften"
ident_attrib_name = "bauteilKlassifikation"


class ExcelIterator(type):
    def __iter__(cls: ExcelBlock) -> Iterator[ExcelBlock]:
        return iter(cls._registry)


class ExcelBlock(metaclass=ExcelIterator):
    _registry: list[ExcelBlock] = list()

    def __init__(self, base_cell: Cell, sheet: Worksheet):
        self._registry.append(self)
        self.base_cell = base_cell
        self.sheet = sheet
        self.pset: classes.PropertySet | None = None
        self.object: classes.Object | None = None

    @property
    def ident_value(self) -> str | None:
        cell = self.sheet.cell(row=self.base_cell.row,
                               column=self.base_cell.column + 2)
        return cell.value

    @property
    def name(self) -> str:
        cell = self.sheet.cell(row=self.base_cell.row,
                               column=self.base_cell.column + 1)
        return cell.value

    @property
    def abbreviation(self) -> str:
        cell = self.sheet.cell(row=self.base_cell.row + 1,
                               column=self.base_cell.column + 1)
        return cell.value.upper()

    @property
    def is_predefined_pset(self) -> bool:
        if self.ident_value is None:
            return True
        else:
            return False

    @property
    def entry(self) -> Cell:
        if self.is_predefined_pset:
            return self.sheet.cell(row=self.base_cell.row + 4, column=self.base_cell.column)
        else:
            return self.sheet.cell(row=self.base_cell.row + 5, column=self.base_cell.column)

    @property
    def aggregates(self) -> set[ExcelBlock] | None:
        if self.is_predefined_pset:
            return None

        child_cell = self.sheet.cell(row=self.base_cell.row + 3, column=self.base_cell.column + 1)
        child_string_list = split_string(child_cell.value)
        excel_block_dict: dict[str, ExcelBlock] = {block.abbreviation: block for block in ExcelBlock}

        children = set()
        for abbrev in child_string_list:
            if abbrev != "-":
                block = find_by_abbreviation(abbrev)
                if block is not None:
                    children.add(block)
                else:
                    logging.error(f"[{self.name}] abbreviation '{abbrev}' doesn't exist!")
        return children

    @property
    def parent_classes(self) -> set[ExcelBlock] | None:
        parents = set()

        parent_text = self.sheet.cell(self.base_cell.row + 2, self.base_cell.column + 1).value
        for abbrev in split_string(parent_text):
            if abbrev != "-":
                parent_block = find_by_abbreviation(abbrev)
                if parent_block is not None:
                    if parent_block.name != ident_pset_name:
                        parents.add(parent_block)
                        parents = set.union(parent_block.parent_classes, parents)
                else:

                    logging.warning(
                        f"[{self.name}] Elternklasse: Kürzel {abbrev.upper()} existiert nicht!")
        return parents

    @property
    def inherits(self) -> set[ExcelBlock]:

        def is_child(value) -> bool:
            last_num = value.split(".")[-1]

            if not last_num.isdigit():
                return False
            last_num = int(last_num)
            if last_num >= 100 and last_num % 100 == 0:
                return True
            else:
                return False

        inherit_blocks = set()

        if self.is_predefined_pset:
            return inherit_blocks

        for block in ExcelBlock:
            if not block.is_predefined_pset:
                parent_txt = block.ident_value.split(".")[:-1]
                parent_txt = ".".join(parent_txt)
                if parent_txt == self.ident_value and is_child(block.ident_value):
                    inherit_blocks.add(block)

        return inherit_blocks

    def create_attributes(self) -> list[classes.Attribute]:
        """
           Iterate over Attributes
           Create Them and find special Datatypes
           """

        if self.is_predefined_pset:
            row = self.base_cell.row + 4
        else:
            row = self.base_cell.row + 5

        entry = self.sheet.cell(row=row, column=self.base_cell.column)

        cell_list = [block.base_cell for block in ExcelBlock]
        attributes: list[classes.Attribute] = list()

        while entry.value is not None and entry not in cell_list and entry.value != "-":
            data_type_text = self.sheet.cell(row=entry.row, column=entry.column + 2).value
            data_type, special = transform_value_types(data_type_text)
            attribute_name = entry.value
            alternative_name = self.sheet.cell(row=entry.row, column=entry.column + 1).value
            attribute = classes.Attribute(self.pset, attribute_name, [""], constants.VALUE_TYPE_LOOKUP[constants.LIST],
                                          data_type=data_type)
            if alternative_name and alternative_name is not None:
                attribute.revit_name = alternative_name
            attributes.append(attribute)

            if special:
                logging.info(
                    f"[{entry.value}] Property: {self.name}:{attribute_name} datatype '{data_type_text}' unbekannt")
                pass

            entry = self.sheet.cell(row=entry.row + 1, column=entry.column)

        return attributes

    def create_predefined_pset(self) -> None:
        self.pset = classes.PropertySet(self.name)
        self.pset.attributes = self.create_attributes()

    def create_object(self) -> classes.Object:
        self.pset = classes.PropertySet(self.name)
        self.pset.attributes = self.create_attributes()
        predef_psets: dict[str, ExcelBlock] = {block.name: block for block in ExcelBlock if block.is_predefined_pset}
        parent_pset = predef_psets.get(ident_pset_name)

        ident_pset = parent_pset.pset.create_child(ident_pset_name)
        ident_attrib = ident_pset.get_attribute_by_name(ident_attrib_name)
        ident_attrib.value = [self.ident_value]
        obj = classes.Object(self.name, ident_attrib)
        obj.add_property_set(self.pset)
        obj.add_property_set(ident_pset)
        obj.ifc_mapping = self.ifc_mapping()
        return obj

    def ifc_mapping(self) -> set[str]:
        if self.is_predefined_pset:
            return set()

        cell = self.sheet.cell(row=self.base_cell.row + 4, column=self.base_cell.column+2)
        value:str = cell.value
        if value is None:
            logging.error(f"[{self.name}]: no IFC Mapping")
            return {"IfcBuildingElementProxy"}

        string_list = value.split("/")
        string_list = [string.strip() for string in string_list]

        for string in string_list:
            if string not in IFC_4_1:
                logging.warning(f"[{self.name}]: '{string}' not in IFC 4.1 Specification")
        return set(string_list)


def transform_value_types(value: str) -> (str, bool):
    special = False
    if value is not None:
        if value.lower() in ["string", "str"]:
            data_type = constants.XS_STRING
        elif value.lower() in ["double"]:
            data_type = constants.XS_DOUBLE
        elif value.lower() in ["boolean", "bool"]:
            data_type = constants.XS_BOOL
        elif value.lower() in ["int", "integer"]:
            data_type = constants.XS_INT
        else:
            special = True
            data_type = constants.XS_STRING
    else:
        data_type = constants.XS_STRING

    return data_type, special


def split_string(text: str) -> list[str] | None:
    if text is None:
        return []
    text = text.split(";")
    for i, item in enumerate(text):
        if "(" in item:
            item = item.split("(")
            text[i] = item[0]
        text[i] = text[i].strip()

    return text


def find_by_abbreviation(abbreviation: str) -> ExcelBlock | None:
    excel_block_dict: dict[str, ExcelBlock] = {block.abbreviation.upper(): block for block in ExcelBlock}
    return excel_block_dict.get(abbreviation.upper())


def start(main_window, path: str) -> None:
    # TODO: add request for Identification Attribute

    def create_blocks() -> None:
        """create Excel Blocks"""
        excel_blocks = set()

        row: tuple[Cell]
        for row in sheet:
            for cell in row:
                if cell.value is not None:
                    text = cell.value.strip()
                    if text in ["name", "name:"]:
                        if sheet.cell(cell.row + 1, cell.column).value == "Kürzel":
                            excel_blocks.add(ExcelBlock(cell, sheet))
                        else:
                            logging.error(f"{sheet.cell(cell.row + 1, cell.column)} hat den Wert 'name'")

    def create_items() -> None:
        """create Objects and PropertySets"""
        predef_psets = [block for block in ExcelBlock if block.is_predefined_pset]
        objects = [block for block in ExcelBlock if not block.is_predefined_pset]

        for predef_pset_block in predef_psets:
            predef_pset_block.create_predefined_pset()

        for object_block in objects:
            obj = object_block.create_object()
            object_block.object = obj

        for block in objects:
            for pset in [block.pset for block in block.parent_classes]:
                new_pset = pset.create_child(pset.name)
                block.object.add_property_set(new_pset)

    def build_object_tree() -> None:
        tree_dict: dict[str, classes.Object] = {obj.ident_attrib.value[0]: obj for obj in classes.Object}

        for ident, item in tree_dict.items():
            ident_list = ident.split(".")[:-1]
            parent_obj = tree_dict.get(".".join(ident_list))

            if parent_obj is not None:
                parent_obj.add_child(item)

        main_window.fill_tree()

    def build_aggregations(gw: graphs_window.GraphWindow) -> None:
        def get_root_blocks() -> set[ExcelBlock]:
            children = set()
            for e_block in ExcelBlock:
                if not e_block.is_predefined_pset:
                    children = set.union(e_block.aggregates, children)
                    children = set.union(e_block.inherits, children)

            r_blocks = set()
            for e_block in ExcelBlock:
                if e_block not in children and not e_block.is_predefined_pset:
                    r_blocks.add(e_block)

            return r_blocks

        def link_child_nodes(node: graphs_window.Node, block: ExcelBlock) -> None:
            aggregate_list = block.aggregates
            inherit_list = block.inherits

            for aggregate_block in aggregate_list:
                if aggregate_block.name != block.name:
                    child_node = graphs_window.Node(aggregate_block.object, gw)
                    relationship = constants.AGGREGATION
                    if aggregate_block in inherit_list:
                        relationship += constants.INHERITANCE

                    node.add_child(child_node, relationship)
                    link_child_nodes(child_node, aggregate_block)
                else:
                    logging.warning(f"[{node.name}] recursive aggregation")

            for inherit_block in inherit_list:
                if inherit_block not in aggregate_list:
                    child_node = graphs_window.Node(inherit_block.object, gw)
                    node.add_child(child_node, constants.INHERITANCE)
                    link_child_nodes(child_node, inherit_block)



        root_blocks = get_root_blocks()

        for block in root_blocks:
            node = graphs_window.Node(block.object, gw)
            scene = graphs_window.AggregationScene(node)
            gw.scenes.append(scene)
            link_child_nodes(node, block)

        gw.update_combo_list()
        for node in gw.root_nodes:
            gw.change_scene(node)
        gw.combo_box.setCurrentIndex(0)
        gw.combo_change()

    with tempfile.TemporaryDirectory() as tmpdirname:

        new_path = os.path.join(tmpdirname,"excel.xlsx")
        shutil.copy2(path,new_path)
        book = openpyxl.load_workbook(new_path)
        sheet = book.active

        create_blocks()
        create_items()

        build_object_tree()
        main_window.graph_window = graphs_window.GraphWindow(main_window, False)
        build_aggregations(main_window.graph_window)
