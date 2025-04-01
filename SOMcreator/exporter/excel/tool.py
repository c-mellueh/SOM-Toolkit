from __future__ import annotations
from typing import TYPE_CHECKING, Callable

import os.path

from openpyxl import Workbook
from openpyxl import styles
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet
import SOMcreator

IDENT_PSET_NAME = "Allgemeine Eigenschaften"
IDENT_PROPERTY_NAME = "bauteilKlassifikation"
NAME = "name"
CLASSES = "classes"
TABLE_STYLE = "TableStyleLight1"
OPTIONAL_FONT = styles.Font(color="4e6ec0")
HEADER_ROW_COUNT = 4
HEADER_COLUMN_COUNT = 5

if TYPE_CHECKING:
    from SOMcreator.exporter.excel import ExcelProperties


class ExportExcel:
    @classmethod
    def get_properties(cls) -> ExcelProperties:
        return SOMcreator.ExcelProperties

    @classmethod
    def get_project(cls) -> SOMcreator.SOMProject:
        return cls.get_properties().project

    @classmethod
    def set_project(cls, project: SOMcreator.SOMProject):
        cls.get_properties().project = project

    @classmethod
    def set_ident_values(cls, pset_name: str, property_name: str):
        cls.get_properties().ident_pset_name = pset_name
        cls.get_properties().ident_property_name = property_name

    @classmethod
    def get_ident_pset_name(cls) -> str:
        if cls.get_properties().ident_pset_name is None:
            pset_name, _ = cls.get_project().get_main_property()
            cls.get_properties().ident_pset_name = pset_name
        return cls.get_properties().ident_pset_name

    @classmethod
    def get_ident_property_name(cls) -> str:
        if cls.get_properties().ident_property_name is None:
            _, property_name = cls.get_project().get_main_property()
            cls.get_properties().ident_property_name = property_name
        return cls.get_properties().ident_property_name

    @classmethod
    def get_class_data(cls, data_dict):
        return data_dict[NAME], data_dict[CLASSES]

    @classmethod
    def create_workbook(cls):
        return Workbook()

    @classmethod
    def directory_of_path_exists(cls, path):
        return os.path.exists(os.path.dirname(path))

    @classmethod
    def _get_name(cls, som_class: SOMcreator.SOMClass):
        return som_class.name

    @classmethod
    def _get_identifier(cls, som_class: SOMcreator.SOMClass):
        return som_class.ident_value or ""

    @classmethod
    def _get_abbreviation(cls, som_class: SOMcreator.SOMClass):
        return som_class.abbreviation or ""

    @classmethod
    def _get_ifc_mapping(cls, som_class: SOMcreator.SOMClass):
        return ";".join(som_class.ifc_mapping) or ""

    @classmethod
    def get_root_parent(cls,som_class:SOMcreator.SOMClass) ->SOMcreator.SOMClass:
        if not som_class.is_child:
            return som_class
        parent = som_class.parent
        while parent.is_child:
            parent = parent.parent
        return parent

    @classmethod
    def fill_main_sheet(cls, sheet: Worksheet,class_list:list[SOMcreator.SOMClass],cell_dict:dict) -> None:
        sheet.title = "Uebersicht"
        titles = ["Kategorie","Name", "Identifier", "Abkürzung", "IfcMapping"]
        getter_functions: list[Callable] = [
            lambda c:cls.get_root_parent(c).name,
            cls._get_name,
            cls._get_identifier,
            cls._get_abbreviation,
            cls._get_ifc_mapping,
        ]
        for column, text in enumerate(titles, start=1):
            sheet.cell(1, column).value = text
        row = 1
        for row, som_class in enumerate(class_list, start=2):
            for column, getter_function in enumerate(getter_functions, start=1):
                sheet.cell(row, column).value = getter_function(som_class)
                if som_class.is_optional(ignore_hirarchy=False):
                    sheet.cell(row, column).font = OPTIONAL_FONT
            class_ws,table_range = cell_dict[som_class]
            hyperlink_text = f"#'{class_ws.title}'!{table_range}"
            sheet.cell(row,2).hyperlink = hyperlink_text
            sheet.cell(row,2).style = "Hyperlink"

        table_range = (
            f"{sheet.cell(1, 1).coordinate}:{sheet.cell(row, len(titles)).coordinate}"
        )
        table = Table(displayName="Uebersicht", ref=table_range)
        style = TableStyleInfo(
            name=TABLE_STYLE,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        sheet.add_table(table)
        cls.autoadjust_column_widths(sheet)

    @classmethod
    def filter_to_sheets(cls, class_list: list[SOMcreator.SOMClass]) -> dict:
        
        root_classes = [c for c in class_list if not c.is_child]
        d = {c.name:{NAME:c.name,CLASSES:[c]} for c in root_classes}
        for som_class in class_list:
            if som_class in root_classes:
                continue
            parent = cls.get_root_parent(som_class)
            d[parent.name][CLASSES].append(som_class)


        
        # d = {
        #     som_class.ident_value: {NAME: som_class.name, CLASSES: []}
        #     for som_class in class_list
        #     if len(som_class.ident_value.split(".")) == 1
        # }

        # for som_class in class_list:
        #     group = som_class.ident_value.split(".")[0]
        #     d[group][CLASSES].append(som_class)
        # d["son"] = {NAME: "Sonstige", CLASSES: []}
        # for group_name, group in list(d.items()):
        #     classes = group[CLASSES]
        #     if len(classes) < 3:
        #         d["son"][CLASSES] += classes
        #         del d[group_name]
        return d

    @classmethod
    def create_class_entry(
        cls, som_class: SOMcreator.SOMClass, sheet, start_row, start_column, table_index
    ):
        if som_class.is_optional(ignore_hirarchy=False):
            font_style = OPTIONAL_FONT
        else:
            font_style = styles.Font()
        sheet.cell(start_row, start_column).value = "bauteilName"
        sheet.cell(start_row, start_column + 1).value = som_class.name

        sheet.cell(start_row + 1, start_column).value = "bauteilKlassifikation"
        sheet.cell(start_row + 1, start_column + 1).value = som_class.ident_value

        sheet.cell(start_row + 2, start_column).value = "Kürzel"
        sheet.cell(start_row + 2, start_column + 1).value = str(som_class.abbreviation)

        sheet.cell(start_row + 3, start_column).value = "Property"
        sheet.cell(start_row + 3, start_column + 1).value = "Propertyset"
        sheet.cell(start_row + 3, start_column + 2).value = "Beispiele / Beschreibung"
        sheet.cell(start_row + 3, start_column + 3).value = "Datentyp"
        sheet.cell(start_row + 3, start_column + 4).value = "Werte"

        for i in range(0, HEADER_ROW_COUNT):
            for k in range(0, HEADER_COLUMN_COUNT):
                sheet.cell(start_row + i, start_column + k).font = font_style
        cls.draw_border(
            sheet,
            [start_row, start_row + 2],
            [start_column, start_column + HEADER_COLUMN_COUNT - 1],
        )
        cls.fill_grey(
            sheet,
            [start_row, start_row + 2],
            [start_column, start_column + HEADER_COLUMN_COUNT - 1],
        )

        pset_start_row = start_row + HEADER_ROW_COUNT
        index = 0
        for property_set in sorted(som_class.get_property_sets(filter=True)):
            for som_property in sorted(property_set.get_properties(filter=True)):
                sheet.cell(pset_start_row + index, start_column).value = (
                    som_property.name
                )
                sheet.cell(pset_start_row + index, start_column + 1).value = (
                    property_set.name
                )
                sheet.cell(pset_start_row + index, start_column + 2).value = (
                    som_property.description
                )
                sheet.cell(pset_start_row + index, start_column + 3).value = (
                    som_property.data_type
                )
                sheet.cell(pset_start_row + index, start_column + 4).value = ";".join(
                    [str(v) for v in som_property.allowed_values]
                )
                if som_property.is_optional(ignore_hirarchy=False):
                    for c in range(HEADER_COLUMN_COUNT):
                        sheet.cell(pset_start_row + index, start_column + c).font = (
                            OPTIONAL_FONT
                        )
                index += 1

        table_start = sheet.cell(pset_start_row - 1, start_column).coordinate
        table_end = sheet.cell(
            pset_start_row + index - 1, start_column + HEADER_COLUMN_COUNT - 1
        ).coordinate
        table_range = f"{table_start}:{table_end}"
        table = Table(
            displayName=f"Tabelle_{str(table_index).zfill(5)}", ref=table_range
        )
        style = TableStyleInfo(
            name=TABLE_STYLE,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        sheet.add_table(table)
        return table_range
    @classmethod
    def draw_border(
        cls, sheet: Worksheet, row_range: [int, int], column_range: [int, int]
    ):
        for row in range(row_range[0], row_range[1] + 1):
            for column in range(column_range[0], column_range[1] + 1):
                left_side = styles.Side(border_style="none", color="FF000000")
                right_side = styles.Side(border_style="none", color="FF000000")
                top_side = styles.Side(border_style="none", color="FF000000")
                bottom_side = styles.Side(border_style="none", color="FF000000")
                if column == column_range[0]:
                    left_side = styles.Side(border_style="thick", color="FF000000")

                if column == column_range[1]:
                    right_side = styles.Side(border_style="thick", color="FF000000")

                if row == row_range[0]:
                    top_side = styles.Side(border_style="thick", color="FF000000")
                if row == row_range[1]:
                    bottom_side = styles.Side(border_style="thick", color="FF000000")
                sheet.cell(row, column).border = styles.Border(
                    left=left_side, right=right_side, top=top_side, bottom=bottom_side
                )

    @classmethod
    def fill_grey(
        cls, sheet: Worksheet, row_range: [int, int], column_range: [int, int]
    ):
        fill = styles.PatternFill(fill_type="solid", start_color="d9d9d9")
        for row in range(row_range[0], row_range[1] + 1):
            for column in range(column_range[0], column_range[1] + 1):
                sheet.cell(row, column).fill = fill

    @classmethod
    def autoadjust_column_widths(cls, sheet: Worksheet, extra_width=0) -> None:
        for i in range(len(list(sheet.columns))):
            column_letter = get_column_letter(i + 1)
            column = sheet[column_letter]
            width = (
                max(
                    [len(cell.value) for cell in column if cell.value is not None],
                    default=2,
                )
                + extra_width
            )
            sheet.column_dimensions[column_letter].width = width

    @classmethod
    def create_property_table(
        cls, classes: list[SOMcreator.SOMClass], sheet: Worksheet
    ):
        property_sets: dict[str, dict[str, int]] = dict()
        sheet.cell(1, 1).value = "PropertySet"
        sheet.cell(1, 2).value = "Property"

        for som_class in classes:
            for pset in som_class.get_property_sets(filter=True):
                if pset.name not in property_sets:
                    property_sets[pset.name] = dict()
                for som_property in pset.get_properties(filter=True):
                    if som_property.name not in property_sets[pset.name]:
                        property_sets[pset.name][som_property.name] = None

        row_index = 2
        for pset_name, property_list in sorted(property_sets.items()):
            for property_name in sorted(property_list.keys()):
                sheet.cell(row_index, 1).value = pset_name
                sheet.cell(row_index, 2).value = property_name
                property_sets[pset_name][property_name] = row_index
                row_index += 1

        col_index = 3
        for row in range(2, row_index - 1):
            for col in range(len(classes)):
                sheet.cell(row + 1, col + 3).value = "✖️"

        for col_index, som_class in enumerate(
            sorted(classes, key=lambda o: o.name), start=3
        ):
            sheet.cell(1, col_index).value = som_class.name
            for pset in som_class.get_property_sets(filter=True):
                for som_property in pset.get_properties(filter=True):
                    row = property_sets[pset.name][som_property.name]
                    sheet.cell(row, col_index).value = "✔️"

        table_range = f"{sheet.cell(1, 1).coordinate}:{sheet.cell(row_index-1, col_index).coordinate}"
        table = Table(displayName="Property_Mapping", ref=table_range)
        style = TableStyleInfo(
            name=TABLE_STYLE,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        sheet.add_table(table)
        cls.autoadjust_column_widths(sheet, 5)
