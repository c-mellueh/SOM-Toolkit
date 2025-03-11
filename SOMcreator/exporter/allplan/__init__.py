import logging

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

import SOMcreator
from SOMcreator.constants import value_constants

TITLES = ["Definition", "Zuweisung", "Mapping"]
COLUMNS = [
    "AttributeName",
    "AttributeTyp",
    "AttributeValue",
    "AttributMin",
    "AttributMax",
    "AttrEinh",
    "AttrEingab",
    "AttVorgabe_I",
    "AttVorgabe_II",
    "AttVorgabe_III",
    "AttVorgabe_IV",
]

INTERNAL_COLUMNS = ["Objekt", "AttributAllplan", "AttributIfc", "Pset", "Type"]


def create_mapping(
    project: SOMcreator.SOMProject, path: str, allplan_mapping_name: str
):
    def transform_datatype(data_type: str) -> str:
        if data_type == value_constants.INTEGER:
            return "Ganzzahl"
        if data_type == value_constants.REAL:
            return "Fließkommazahl"
        return "Text"

    def create_definition(worksheet: Worksheet) -> dict[str, str]:

        for x, text in enumerate(COLUMNS):
            worksheet.cell(row=1, column=x + 1, value=text)

        property_dict: dict[str, str] = dict()

        for som_property in project.get_properties(filter=True):
            data_type = som_property.data_type

            if som_property.name in property_dict:
                if data_type != property_dict[som_property.name]:
                    if not som_property.property_set or not som_property.property_set.som_class:
                        continue
                    logging.info(
                        f"Achtung bei {som_property.property_set.som_class.name} -> {som_property.property_set.name}"
                        f":{som_property.name} neuer Datentyp: {data_type} "
                        f" alter Datentyp: {property_dict[som_property.name]}"
                    )
            else:
                property_dict[som_property.name] = data_type

        for row_index, [name, data_type] in enumerate(property_dict.items()):
            row = 2 + row_index
            worksheet.cell(row=row, column=1, value=name)
            worksheet.cell(row=row, column=2, value=transform_datatype(data_type))
            if data_type == value_constants.BOOLEAN:
                worksheet.cell(row=row, column=7, value="CheckBox")
        return property_dict

    def create_zuweisung(kenner: str, worksheet: Worksheet):

        def get_attrib_count(som_class: SOMcreator.SOMClass):
            return sum(
                len([attrib for attrib in pset.get_properties(filter=True)])
                for pset in som_class.get_property_sets(filter=True)
            )

        max_attribs = max(
            get_attrib_count(som_class) for som_class in project.get_classes(filter=True)
        )
        header = ["Kenner"] + ["Wert", "Name"] * max_attribs
        [
            worksheet.cell(1, i + 1, text) for i, text in enumerate(header)
        ]  # print Header
        worksheet.cell(2, 1, kenner)
        row_index = 2
        for som_class in project.get_classes(filter=True):
            worksheet.cell(row_index, 2, som_class.ident_value)
            col_index = 3
            for propery_set in som_class.get_property_sets(filter=True):
                for som_property in propery_set.get_properties(filter=True):
                    if som_property.name != kenner:
                        worksheet.cell(row_index, col_index, som_property.name)
                        col_index += 2

            row_index += 1

    def create_internal_mapping(property_dict: dict[str, str], worksheet: Worksheet):
        def transform_type(t: str) -> str:
            if t == value_constants.INTEGER:
                return "IfcInteger"
            if t == value_constants.REAL:
                return "IfcReal"
            if t == value_constants.BOOLEAN:
                return "IfcBoolean"
            return "IfcLabel"

        for x, text in enumerate(INTERNAL_COLUMNS):
            worksheet.cell(row=1, column=x + 1, value=text)
        worksheet.cell(2, 1, "All")
        for row_index, (property_name, property_datatype) in enumerate(
            sorted(property_dict.items())
        ):
            worksheet.cell(2 + row_index, 2, property_name)
            worksheet.cell(2 + row_index, 4, allplan_mapping_name)
            worksheet.cell(2 + row_index, 5, transform_type(property_datatype))

    wb = Workbook()
    ws = wb.active
    if ws is None:
        return
    ws.title = TITLES[0]
    ad = create_definition(ws)
    create_zuweisung(
        "bauteilKlassifikation", wb.create_sheet(TITLES[1])
    )  # Todo: Make bk variable
    create_internal_mapping(ad, wb.create_sheet(TITLES[2]))
    wb.save(path)
