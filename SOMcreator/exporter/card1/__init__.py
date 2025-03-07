import logging

from openpyxl import load_workbook, Workbook
import SOMcreator


def create_mapping(
    src_path: str, dest_path: str, project: SOMcreator.SOMProject
) -> None:
    def _create_sheet(obj: SOMcreator.SOMClass, workbook: Workbook, name):
        new_sheet = workbook.create_sheet(name)
        properties = set()
        for property_set in obj.get_property_sets(filter=True):
            for som_property in property_set.get_properties(filter=True):
                properties.add(som_property.name)

        for i, property_name in enumerate(sorted(properties), start=1):
            new_sheet.cell(1, i).value = property_name

    export_wb = Workbook()
    export_wb.active.title = "Hilfe"
    wb = load_workbook(src_path)
    sheet = wb.active
    important_rows = [
        row for i, row in enumerate(sheet.rows) if row[2].value is not None and i != 0
    ]
    object_dict = {
        som_class.identifier_property.allowed_values[0]: som_class
        for som_class in project.get_classes(filter=True)
        if not som_class.is_concept
    }

    for row in important_rows:
        bauteil_bez_card, bauteil_bez_2, bauteilklass = map(lambda x: x.value, row)
        obj = object_dict.get(bauteilklass)
        if obj is None:
            logging.warning(f"identifier '{bauteilklass}' not found")
            continue
        _create_sheet(obj, export_wb, bauteil_bez_card)
    export_wb.save(dest_path)
