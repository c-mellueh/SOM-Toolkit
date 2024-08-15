from __future__ import annotations
from typing import TYPE_CHECKING

import os.path

from openpyxl import Workbook
from openpyxl import styles
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

import SOMcreator
from SOMcreator import classes

IDENT_PSET_NAME = "Allgemeine Eigenschaften"
IDENT_ATTRIB_NAME = "bauteilKlassifikation"
NAME = "name"
OBJECTS = "objects"
TABLE_STYLE = "TableStyleLight1"
OPTIONAL_FONT = styles.Font(color="4e6ec0")

if TYPE_CHECKING:
    from SOMcreator.module.export_excel import ExportExcelProperties


class ExportExcel:
    @classmethod
    def get_properties(cls) -> ExportExcelProperties:
        return SOMcreator.ExportExcelProperties

    @classmethod
    def get_project(cls) -> SOMcreator.Project:
        return cls.get_properties().project

    @classmethod
    def set_project(cls, project: SOMcreator.Project):
        cls.get_properties().project = project

    @classmethod
    def set_ident_values(cls, pset_name: str, attribute_name: str):
        cls.get_properties().ident_pset_name = pset_name
        cls.get_properties().ident_attribute_name = attribute_name

    @classmethod
    def get_ident_pset_name(cls) -> str:
        if cls.get_properties().ident_pset_name is None:
            pset_name, attribute_name = cls.get_project().get_main_attribute()
            cls.get_properties().ident_pset_name = pset_name
        return cls.get_properties().ident_pset_name

    @classmethod
    def get_ident_attribute_name(cls) -> str:
        if cls.get_properties().ident_attribute_name is None:
            pset_name, attribute_name = cls.get_project().get_main_attribute()
            cls.get_properties().ident_attribute_name = attribute_name
        return cls.get_properties().ident_attribute_name

    @classmethod
    def get_object_data(cls, data_dict):
        return data_dict[NAME], data_dict[OBJECTS]

    @classmethod
    def create_workbook(cls):
        return Workbook()

    @classmethod
    def directory_of_path_exists(cls, path):
        return os.path.exists(os.path.dirname(path))

    @classmethod
    def _get_name(cls, obj: classes.Object):
        return obj.name

    @classmethod
    def _get_identifier(cls, obj: classes.Object):
        return obj.ident_value or ""

    @classmethod
    def _get_abbreviation(cls, obj: classes.Object):
        return obj.abbreviation or ""

    @classmethod
    def _get_ifc_mapping(cls, obj: classes.Object):
        return ";".join(obj.ifc_mapping) or ""

    @classmethod
    def fill_main_sheet(cls, sheet: Worksheet) -> None:
        project = cls.get_project()
        sheet.title = "Uebersicht"
        titles = ["bauteilName", "bauteilKlassifikation", "abkuerzung", "IfcMapping"]
        getter_functions = [cls._get_name, cls._get_identifier, cls._get_abbreviation, cls._get_ifc_mapping]
        for column, text in enumerate(titles, start=1):
            sheet.cell(1, column).value = text

        row = 1
        for row, obj in enumerate(sorted(project.objects), start=2):
            for column, getter_function in enumerate(getter_functions, start=1):
                sheet.cell(row, column).value = getter_function(obj)
                if obj.optional:
                    sheet.cell(row, column).font = OPTIONAL_FONT

        table_range = f"{sheet.cell(1, 1).coordinate}:{sheet.cell(row, len(titles)).coordinate}"
        table = Table(displayName="Uebersicht", ref=table_range)
        style = TableStyleInfo(name=TABLE_STYLE, showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        sheet.add_table(table)
        cls.autoadjust_column_widths(sheet)

    @classmethod
    def filter_to_sheets(cls, ) -> dict:
        project = cls.get_project()
        d = {obj.ident_value: {NAME: obj.name, OBJECTS: []} for obj in project.objects if
             len(obj.ident_value.split(".")) == 1}
        for obj in project.objects:
            group = obj.ident_value.split(".")[0]
            d[group][OBJECTS].append(obj)
        d["son"] = {NAME: "Sonstige", OBJECTS: []}
        for group_name, group in list(d.items()):
            objects = group[OBJECTS]
            if len(objects) < 3:
                d["son"][OBJECTS] += objects
                del d[group_name]
        return d

    @classmethod
    def create_object_entry(cls, obj: classes.Object, sheet, start_row, start_column, table_index):
        if obj.optional:
            font_style = OPTIONAL_FONT
        else:
            font_style = styles.Font()

        sheet.cell(start_row, start_column).value = "bauteilName"
        sheet.cell(start_row, start_column + 1).value = obj.name

        sheet.cell(start_row + 1, start_column).value = "bauteilKlassifikation"
        sheet.cell(start_row + 1, start_column + 1).value = obj.ident_value

        sheet.cell(start_row + 2, start_column).value = "Kürzel"
        sheet.cell(start_row + 2, start_column + 1).value = str(obj.abbreviation)

        sheet.cell(start_row + 3, start_column).value = "Property"
        sheet.cell(start_row + 3, start_column + 1).value = "Propertyset"
        sheet.cell(start_row + 3, start_column + 2).value = "Beispiele / Beschreibung"
        sheet.cell(start_row + 3, start_column + 3).value = "Datentyp"

        for i in range(0, 4):
            for k in range(0, 4):
                sheet.cell(start_row + i, start_column + k).font = font_style
        cls.draw_border(sheet, [start_row, start_row + 2], [start_column, start_column + 4])
        cls.fill_grey(sheet, [start_row, start_row + 2], [start_column, start_column + 3])

        pset_start_row = start_row + 4
        index = 0
        for property_set in sorted(obj.property_sets):
            for attribute in sorted(property_set.attributes):
                sheet.cell(pset_start_row + index, start_column).value = attribute.name
                sheet.cell(pset_start_row + index, start_column + 1).value = property_set.name
                sheet.cell(pset_start_row + index, start_column + 2).value = attribute.description
                sheet.cell(pset_start_row + index, start_column + 3).value = attribute.data_type
                if attribute.optional:
                    sheet.cell(pset_start_row + index, start_column).font = OPTIONAL_FONT
                    sheet.cell(pset_start_row + index, start_column + 1).font = OPTIONAL_FONT
                    sheet.cell(pset_start_row + index, start_column + 2).font = OPTIONAL_FONT
                index += 1

        table_start = sheet.cell(pset_start_row - 1, start_column).coordinate
        table_end = sheet.cell(pset_start_row + index - 1, start_column + 3).coordinate
        table_range = f"{table_start}:{table_end}"
        table = Table(displayName=f"Tabelle_{str(table_index).zfill(5)}", ref=table_range)
        style = TableStyleInfo(name=TABLE_STYLE, showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        sheet.add_table(table)

    @classmethod
    def draw_border(cls, sheet: Worksheet, row_range: [int, int], column_range: [int, int]):
        for row in range(row_range[0], row_range[1] + 1):
            for column in range(column_range[0], column_range[1] + 1):
                left_side = styles.Side(border_style="none", color="FF000000")
                right_side = styles.Side(border_style="none", color="FF000000")
                top_side = styles.Side(border_style="none", color="FF000000")
                bottom_side = styles.Side(border_style="none", color="FF000000")
                if column == column_range[0]:
                    left_side = styles.Side(border_style="thick", color="FF000000")

                if column == column_range[1]:
                    right_side = styles.Side(border_style="thick", color="FF000000")

                if row == row_range[0]:
                    top_side = styles.Side(border_style="thick", color="FF000000")
                if row == row_range[1]:
                    bottom_side = styles.Side(border_style="thick", color="FF000000")
                sheet.cell(row, column).border = styles.Border(left=left_side, right=right_side, top=top_side,
                                                               bottom=bottom_side)

    @classmethod
    def fill_grey(cls, sheet: Worksheet, row_range: [int, int], column_range: [int, int]):
        fill = styles.PatternFill(fill_type="solid", start_color="d9d9d9")
        for row in range(row_range[0], row_range[1] + 1):
            for column in range(column_range[0], column_range[1] + 1):
                sheet.cell(row, column).fill = fill

    @classmethod
    def autoadjust_column_widths(cls, sheet: Worksheet) -> None:
        for i in range(len(list(sheet.columns))):
            column_letter = get_column_letter(i + 1)
            column = sheet[column_letter]
            width = max([len(cell.value) for cell in column if cell.value is not None], default=2)
            sheet.column_dimensions[column_letter].width = width
