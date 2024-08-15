import logging

from openpyxl import load_workbook, Workbook

from .. import classes


def create_mapping(src_path: str, dest_path: str, project: classes.Project) -> None:
    def _create_sheet(obj: classes.Object, workbook: Workbook, name):
        new_sheet = workbook.create_sheet(name)
        attributes = set()
        for property_set in obj.property_sets:
            for attribute in property_set.attributes:
                attributes.add(attribute.name)

        for i, attrib_name in enumerate(sorted(attributes), start=1):
            new_sheet.cell(1, i).value = attrib_name

    export_wb = Workbook()
    export_wb.active.title = "Hilfe"
    wb = load_workbook(src_path)
    sheet = wb.active
    important_rows = [row for i, row in enumerate(sheet.rows) if row[2].value is not None and i != 0]
    object_dict = {obj.ident_attrib.value[0]: obj for obj in project.objects if
                   not obj.is_concept}

    for row in important_rows:
        bauteil_bez_card, bauteil_bez_2, bauteilklass = map(lambda x: x.value, row)
        obj = object_dict.get(bauteilklass)
        if obj is None:
            logging.warning(f"identifier '{bauteilklass}' not found")
            continue
        _create_sheet(obj, export_wb, bauteil_bez_card)
    export_wb.save(dest_path)
