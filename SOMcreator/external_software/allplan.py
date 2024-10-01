import logging

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

import SOMcreator
from ..constants import value_constants

TITLES = ["Definition", "Zuweisung", "Mapping"]
COLUMNS = ["AttributeName",
           "AttributeTyp",
           "AttributeValue",
           "AttributMin",
           "AttributMax",
           "AttrEinh",
           "AttrEingab",
           "AttVorgabe_I",
           "AttVorgabe_II",
           "AttVorgabe_III",
           "AttVorgabe_IV"]

INTERNAL_COLUMNS = ["Objekt", "AttributAllplan", "AttributIfc", "Pset", "Type"]


def create_mapping(project: SOMcreator.Project, path: str, allplan_mapping_name: str):
    def transform_datatype(data_type: str) -> str:
        if data_type == value_constants.INTEGER:
            return "Ganzzahl"
        if data_type == value_constants.REAL:
            return "Fließkommazahl"
        return "Text"

    def create_definition(worksheet: Worksheet) -> dict[str, str]:

        for x, text in enumerate(COLUMNS):
            worksheet.cell(row=1, column=x + 1, value=text)

        attribute_dict: dict[str, str] = dict()

        for attribute in SOMcreator.Attribute:
            data_type = attribute.data_type

            if attribute.name in attribute_dict:
                if data_type != attribute_dict[attribute.name]:
                    logging.warning(f"Achtung bei {attribute.property_set.object.name} -> {attribute.property_set.name}"
                                    f":{attribute.name} neuer Datentyp: {data_type} "
                                    f" alter Datentyp: {attribute_dict[attribute.name]}")
            else:
                attribute_dict[attribute.name] = data_type

        for row_index, [name, data_type] in enumerate(attribute_dict.items()):
            row = 2 + row_index
            worksheet.cell(row=row, column=1, value=name)
            worksheet.cell(row=row, column=2, value=transform_datatype(data_type))
            if data_type == value_constants.BOOLEAN:
                worksheet.cell(row=row, column=7, value="CheckBox")
        return attribute_dict

    def create_zuweisung(kenner: str, worksheet: Worksheet):

        def get_attrib_count(obj: SOMcreator.Object):
            return sum(len([attrib for attrib in pset.get_attributes(filter=True)]) for pset in
                       obj.get_property_sets(filter=True))

        max_attribs = max(
            get_attrib_count(obj) for obj in project.get_objects(filter=True))
        header = ["Kenner"] + ["Wert", "Name"] * max_attribs
        [worksheet.cell(1, i + 1, text) for i, text in enumerate(header)]  # print Header
        worksheet.cell(2, 1, kenner)
        row_index = 2
        for obj in project.get_objects(filter=True):
            worksheet.cell(row_index, 2, obj.ident_value)
            col_index = 3
            for propery_set in obj.get_property_sets(filter=True):
                for attribute in propery_set.get_attributes(filter=True):
                    if attribute.name != kenner:
                        worksheet.cell(row_index, col_index, attribute.name)
                        col_index += 2

            row_index += 1

    def create_internal_mapping(attribute_dict: dict[str, str], worksheet: Worksheet):
        def transform_type(t: str) -> str:
            if t == value_constants.INTEGER:
                return "IfcInteger"
            if t == value_constants.REAL:
                return "IfcReal"
            if t == value_constants.BOOLEAN:
                return "IfcBoolean"
            return "IfcLabel"

        for x, text in enumerate(INTERNAL_COLUMNS):
            worksheet.cell(row=1, column=x + 1, value=text)
        worksheet.cell(2, 1, "All")
        for row_index, (attribute_name, attribute_datatype) in enumerate(sorted(attribute_dict.items())):
            worksheet.cell(2 + row_index, 2, attribute_name)
            worksheet.cell(2 + row_index, 4, allplan_mapping_name)
            worksheet.cell(2 + row_index, 5, transform_type(attribute_datatype))

    wb = Workbook()
    ws = wb.active
    ws.title = TITLES[0]

    ad = create_definition(ws)
    create_zuweisung("bauteilKlassifikation", wb.create_sheet(TITLES[1]))  # Todo: Make bk variable
    create_internal_mapping(ad, wb.create_sheet(TITLES[2]))
    wb.save(path)
