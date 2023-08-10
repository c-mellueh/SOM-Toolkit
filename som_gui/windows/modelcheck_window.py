from __future__ import annotations

import logging
import os
import tempfile
from datetime import datetime
from time import time, sleep
from typing import TYPE_CHECKING

import ifcopenshell
import openpyxl
from PySide6.QtCore import QObject, Signal, QRunnable, QThreadPool,QSize
from PySide6.QtGui import QResizeEvent
from PySide6.QtWidgets import QTableWidget, QTableWidgetItem, QLineEdit
from SOMcreator import classes
from SOMcreator import constants as som_constants
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.worksheet.table import Table, TableStyleInfo

from ..icons import get_icon
from ..ifc_modification import modelcheck, sql, issues
from ..qt_designs import ui_modelcheck
from ..settings import get_ifc_path, get_issue_path, set_ifc_path, set_issue_path
from .ifc_mod_window import IfcWindow

if TYPE_CHECKING:
    from ..main_window import MainWindow

FILE_SPLIT = "; "
HEADER = ["Datum", "GUID", "Beschreibung", "Typ", "Name", "PropertySet", "Attribut", "Datei",
          "Bauteilklassifikation"]


class ModelcheckWindow(IfcWindow):
    def __init__(self, main_window: MainWindow):
        super(ModelcheckWindow, self).__init__(main_window)
        self.setWindowTitle("Modelcheck")
        self.data_base_path = None
        self.widget.line_edit_ident_pset.textEdited.connect(self.fill_table)
        self.widget.line_edit_ident_attribute.textEdited.connect(self.fill_table)
        self.create_table()
        self.fill_table()
        self.set_fixed_sizes()
        self.adjustSize()

    def set_fixed_sizes(self):
        def set_line_edit_size(le:QLineEdit):
            text = le.text()
            fm = le.fontMetrics()
            width = fm.boundingRect(text).width()
            le.setMinimumSize(max(width,le.width()),le.height())

        def set_table_height(table:QTableWidget):
            height = table.horizontalHeader().height()+4
            height+=sum(table.rowHeight(row) for row in range(table.rowCount()))
            for col in range(table.columnCount()):
                print(table.rowHeight(col))

            print(height)
            size = QSize(table.width(),height)
            table.setMinimumSize(size)

        set_line_edit_size(self.widget.line_edit_ident_pset)
        set_line_edit_size(self.widget.line_edit_ident_attribute)
        #set_table_height(self.table_widget)

    def create_table(self):
        self.table_widget = QTableWidget()
        self.widget.verticalLayout.addWidget(self.table_widget)
        if (self.table_widget.columnCount() < 2):
            self.table_widget.setColumnCount(2)
        __qtablewidgetitem = QTableWidgetItem()
        self.table_widget.setHorizontalHeaderItem(0, __qtablewidgetitem)
        __qtablewidgetitem.setText("Fehlertyp")
        __qtablewidgetitem1 = QTableWidgetItem()
        self.table_widget.setHorizontalHeaderItem(1, __qtablewidgetitem1)
        __qtablewidgetitem1.setText("Beschreibung")
        self.table_widget.setObjectName(u"table_widget")
        self.table_widget.horizontalHeader().setStretchLastSection(True)
        self.table_widget.verticalHeader().setVisible(False)
        self.table_widget.show()

    def fill_table(self):
        issues = self.get_issue_description()
        self.table_widget.setRowCount(0)
        self.table_widget.setRowCount(len(issues))
        for index, (key, description) in enumerate(sorted(issues.items())):
            self.table_widget.setItem(index, 0, QTableWidgetItem(f"Fehler {key}"))
            self.table_widget.setItem(index, 1, QTableWidgetItem(description))

    def get_issue_description(self):
        main_pset = self.widget.line_edit_ident_pset.text()
        main_attrib = self.widget.line_edit_ident_attribute.text()

        return {1: f"Propertyset '{main_pset}' fehlt",
                2: f"{main_pset}:{main_attrib} fehlt",
                3: "Bauteilklassifikation nicht in SOM vorhanden",
                4: "GUID ist in mehreren Dateien identisch",
                5: "PropertySet Fehlt",
                6: "Attribut Fehlt",
                7: "Attribut hat falschen Wert",
                8: "Gruppe hat falsches Subelement",
                9: "Zwischenebene besitzt verschiedene Klassen als Subelement",
                10: "Gruppe besitzt keine Subelemente",
                11: "Element hat keine Gruppenzuweisen",
                12: "Zu Viele Subelemente"}

    def start_task(self) -> tuple:
        proj, ifc, pset, attribute, export_path = super(ModelcheckWindow, self).start_task()
        self.data_base_path = tempfile.NamedTemporaryFile().name
        print(f"Database: {self.data_base_path}")

        self.runner = Modelcheck(ifc, proj, pset, attribute, self.data_base_path, export_path)

        self.runner.signaller.started.connect(self.on_started)
        self.runner.signaller.finished.connect(self.on_finished)
        self.runner.signaller.progress.connect(self.update_progress_bar)
        self.runner.signaller.status.connect(self.update_status)
        self.widget.button_close.clicked.connect(self.runner.abort)
        self.thread_pool.start(self.runner)
        return proj,ifc,pset,attribute,export_path

class Modelcheck(QRunnable):
    def __init__(self, path, project, property_set, attribute, db_path, export_path):
        super(Modelcheck, self).__init__()
        self.path = path
        self.project = project
        self.property_set = property_set
        self.attribute = attribute
        self.db_path = db_path
        self.signaller = Signaller()
        self.export_path = export_path
        self.is_aborted = False
        self.object_count: int = 0
        self.checked_objects: int = 0
        self.base_name: str = ""
        self.ident_dict = dict()
        self.group_dict: dict[ifcopenshell.entity_instance] = dict()
        self.group_parent_dict = dict()

    def update_progress(self):
        self.checked_objects += 1
        progress = int(self.checked_objects / self.object_count * 100)
        self.signaller.progress.emit(progress)
        self.signaller.status.emit(f"Check '{self.base_name}' [{self.checked_objects}/{self.object_count}]")

    def get_check_file_list(self) -> set[str]:
        ifc_paths = self.path
        check_list = set()
        if not isinstance(ifc_paths, list):
            if not isinstance(ifc_paths, str):
                return check_list
            if not os.path.isfile(ifc_paths):
                return check_list
            check_list.add(ifc_paths)

        else:
            for path in ifc_paths:
                if os.path.isdir(path):
                    for file in os.listdir(path):
                        file_path = os.path.join(path, file)
                        check_list.add(file_path)
                else:
                    check_list.add(path)

        return set(file for file in check_list if file.lower().endswith(".ifc"))

    def abort(self):
        self.is_aborted = True
        self.signaller.status.emit("Modelcheck abgebrochen")
        self.signaller.progress.emit(0)

    def run(self) -> None:
        self.signaller.started.emit(self.path)
        sql.guids = dict()
        sql.create_tables(self.db_path)

        files = self.get_check_file_list()
        for file in files:
            if self.is_aborted:
                continue
            self.check_file(file)

        if self.is_aborted:
            self.signaller.finished.emit("ABORT")
            return
        self.signaller.finished.emit(self.path)
        self.signaller.status.emit("Modelcheck Done!")
        self.create_issues()

    def check_file(self, file):

        self.base_name = os.path.basename(file)
        self.signaller.status.emit(f"Import {self.base_name}")
        self.signaller.progress.emit(0)
        sleep(0.1)

        ifc = ifcopenshell.open(file)
        self.signaller.status.emit(f"Import Done!")
        sleep(0.1)
        self.check_all_elements(ifc)

    def check_all_elements(self, ifc: ifcopenshell.file):
        def count_dict(d):
            """counts how many elements are in dict"""
            for value in d.values():
                self.object_count += 1
                count_dict(value)

        self.signaller.status.emit(f"Check '{self.base_name}'")
        sql.remove_existing_issues(self.db_path, self.project.name, datetime.today(), self.base_name)

        root_groups: list[ifcopenshell.entity_instance] = [group for group in ifc.by_type("IfcGroup") if
                                                           not modelcheck.get_parent_group(group)]
        self.ident_dict = {obj.ident_value: obj for obj in self.project.objects}
        self.group_dict = dict()
        self.group_parent_dict = dict()

        for element in root_groups:
            self.group_dict[element] = dict()
            modelcheck.build_group_structure(element, self.group_dict[element], self.property_set, self.attribute,
                                             self.group_parent_dict)

        self.checked_objects = 0
        self.object_count = 0
        count_dict(self.group_dict)

        for element in self.group_dict:
            if self.is_aborted:
                return
            self.check_group_structure(element, self.group_dict, 0)

        for element in ifc.by_type("IfcElement"):
            group_assignment = [assignment for assignment in getattr(element, "HasAssignments", []) if
                                assignment.is_a("IfcRelAssignsToGroup")]
            if not group_assignment:
                issues.no_group_issue(self.db_path, element)

    def create_issues(self):
        def get_max_width():
            col_widths = [0 for _ in range(worksheet.max_column)]
            for row_index, row in enumerate(worksheet, start=1):
                for col_index, cell in enumerate(row):
                    length = len(str(cell.value))
                    if length > col_widths[col_index]:
                        col_widths[col_index] = length
            return col_widths

        directory = os.path.dirname(self.export_path)
        if not os.path.exists(directory):
            os.mkdir(directory)

        issues = sql.query_issues(self.db_path)

        self.signaller.status.emit(f"{len(issues)} Fehler gefunden!")

        if len(issues) == 0:
            self.signaller.status.emit("Modelle fehlerfrei!")
            return

        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        for col_index, value in enumerate(HEADER, start=1):
            worksheet.cell(1, col_index, value)

        last_cell = worksheet.cell(1, 8)

        for row_index, column in enumerate(issues, start=2):
            for column_index, value in enumerate(column, start=1):
                last_cell = worksheet.cell(row_index, column_index, value)  # remove Whitespace

        table_zone = f"A1:{last_cell.coordinate}"
        tab = Table(displayName="Issues", ref=table_zone)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False,
                               showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        worksheet.add_table(tab)

        max_widths = get_max_width()

        # autoFit Column
        dim_holder = DimensionHolder(worksheet=worksheet)
        for col in range(worksheet.min_column, worksheet.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(worksheet, min=col, max=col,
                                                                 width=max_widths[col - 1] * 1.1)
        worksheet.column_dimensions = dim_holder

        self.save_workbook(workbook, self.export_path)

    def save_workbook(self, workbook, path):
        try:
            workbook.save(path)
        except PermissionError:
            print("-" * 60)
            print(f"folgende Datei ist noch geöffnet: {path} \n Datei schließen und beliebige Taste Drücken")
            input("Achtung! Datei wird danach überschrieben!")
            self.save_workbook(workbook, path)

    def check_group_structure(self, element: ifcopenshell.entity_instance, group_dict, layer_index: int):

        def check_collector_group():
            sql.db_create_entity(self.db_path, element, self.project.name, self.base_name, identifier)
            for sub_ident in sub_idents:
                if sub_ident != identifier:
                    issues.subgroup_issue(self.db_path, element.GlobalId, sub_ident)

        def check_real_element():
            """checks group or element that is not a collector"""

            def loop_parent(el: classes.Aggregation) -> classes.Aggregation:
                if el.parent_connection != som_constants.INHERITANCE:
                    return el.parent
                else:
                    return loop_parent(el.parent)

            def check_correct_parent():
                parent_group = self.group_parent_dict.get(self.group_parent_dict.get(element))
                allowed_parents = set(loop_parent(aggreg) for aggreg in object_rep.aggregations)
                logging.info(
                    f"Group {element.GlobalId} Allowed parents= {[aggreg.name for aggreg in allowed_parents if aggreg is not None]}")
                if parent_group is None:
                    if not None in allowed_parents:
                        logging.warning(f"Group {element.GlobalId} -> no parent group")
                    return
                parent_object = self.ident_dict.get(
                    modelcheck.get_identifier(parent_group, self.property_set, self.attribute))
                if parent_object is None:
                    logging.warning(f"Group {element.GlobalId} -> no parent obj")
                    return

                parent_in_allowed_parents = bool(parent_object.aggregations.union().intersection(allowed_parents))
                if not (parent_in_allowed_parents or None in allowed_parents):
                    issues.parent_issue(self.db_path, element, parent_group, self.property_set, self.attribute,
                                        element_type)

            def check_repetetive_group():
                """Gruppe besitzt mehrere Subelemente mit der selben Bauteilklassifikation"""
                if len(sub_idents) != len(
                        [modelcheck.get_identifier(sub_group, self.property_set, self.attribute) for sub_group in
                         group_dict[element]]):
                    logging.warning(
                        f"Gruppe {element.GlobalId} -> besitzt mehrere Subelemente mit der selben Bauteilklassifikation")
                    issues.repetetive_group_issue(self.db_path, element)

            modelcheck.check_element(element, self.property_set, self.attribute, self.db_path, self.base_name,
                                     self.ident_dict, element_type, self.project.name)

            object_rep = self.ident_dict.get(identifier)
            if object_rep is None:
                logging.warning(f"Group {element.GlobalId} -> no obj rep")
                return
            check_correct_parent()
            if element_type == modelcheck.GROUP:
                check_repetetive_group()

        self.update_progress()

        if self.is_aborted:
            return

        logging.info(f"Check Element {element.GlobalId}")

        identifier = modelcheck.get_identifier(element, self.property_set, self.attribute)
        logging.info(f"Identifier {identifier}")

        even_layer = layer_index % 2 == 0
        logging.info(f"Even Layer: {even_layer}")

        if element.is_a("IfcElement"):
            element_type = modelcheck.ELEMENT
            check_real_element()
            return
        else:
            element_type = modelcheck.GROUP

        sub_idents = {modelcheck.get_identifier(sub_group, self.property_set, self.attribute) for sub_group in
                      group_dict[element]}
        sub_entities = group_dict[element].keys()

        if even_layer:
            check_collector_group()
        else:
            check_real_element()

        if len(sub_entities) == 0:
            issues.empty_group_issue(self.db_path, element)

        for sub_group in group_dict[element]:
            self.check_group_structure(sub_group, group_dict[element], layer_index + 1)


class Signaller(QObject):
    started = Signal(str)
    finished = Signal(str)
    progress = Signal(int)
    status = Signal(str)
