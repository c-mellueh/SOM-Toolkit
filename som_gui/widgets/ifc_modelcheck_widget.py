from __future__ import annotations

import logging
import os
import tempfile
from datetime import datetime
from typing import TYPE_CHECKING

import ifcopenshell
import openpyxl
from PySide6.QtWidgets import QTableWidget, QTableWidgetItem
from SOMcreator import classes
from SOMcreator.constants import value_constants
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.worksheet.table import Table, TableStyleInfo

from som_gui.ifc_modification import modelcheck, sql, issues
from som_gui.widgets.ifc_widget import IfcWidget, IfcRunner

if TYPE_CHECKING:
    from som_gui.main_window import MainWindow
    from ..windows.modelcheck.modelcheck_window import ModelcheckWindow
FILE_SPLIT = "; "
HEADER = ["Datum", "GUID", "Beschreibung", "Typ", "Name", "PropertySet", "Attribut", "Datei",
          "Bauteilklassifikation"]


def count_dict(prod: dict, c=0):
    for mykey, value in prod.items():
        if isinstance(value, dict):
            # calls repeatedly
            c = count_dict(value, c + 1)
        else:
            c += 1
    return c


class ModelcheckWidget(IfcWidget):
    def __init__(self, main_window: MainWindow):
        super(ModelcheckWidget, self).__init__(main_window)
        self.setWindowTitle("Modelcheck")
        self.data_base_path = None
        self.widget.line_edit_ident_pset.textEdited.connect(self.fill_table)
        self.widget.line_edit_ident_attribute.textEdited.connect(self.fill_table)
        self.table_widget = QTableWidget()
        self.create_table()
        self.fill_table()
        self.adjustSize()
        self.runner: Modelcheck | None = None

    @property
    def ident_attribute(self):
        return self.widget.line_edit_ident_attribute.text()

    @property
    def ident_pset(self):
        return self.widget.line_edit_ident_pset.text()

    def create_table(self):
        self.widget.verticalLayout.addWidget(self.table_widget)
        if self.table_widget.columnCount() < 2:
            self.table_widget.setColumnCount(2)
        __qtablewidgetitem = QTableWidgetItem()
        self.table_widget.setHorizontalHeaderItem(0, __qtablewidgetitem)
        __qtablewidgetitem.setText("Fehlertyp")
        __qtablewidgetitem1 = QTableWidgetItem()
        self.table_widget.setHorizontalHeaderItem(1, __qtablewidgetitem1)
        __qtablewidgetitem1.setText("Beschreibung")
        self.table_widget.setObjectName(u"table_widget")
        self.table_widget.horizontalHeader().setStretchLastSection(True)
        self.table_widget.verticalHeader().setVisible(False)
        self.table_widget.show()

    def fill_table(self):
        _issues = self.get_issue_description()
        self.table_widget.setRowCount(0)
        self.table_widget.setRowCount(len(_issues))
        for index, (key, description) in enumerate(sorted(_issues.items())):
            self.table_widget.setItem(index, 0, QTableWidgetItem(f"Fehler {key}"))
            self.table_widget.setItem(index, 1, QTableWidgetItem(description))

    def get_issue_description(self):
        main_pset = self.widget.line_edit_ident_pset.text()
        main_attrib = self.widget.line_edit_ident_attribute.text()

        return {1: f"Propertyset '{main_pset}' fehlt",
                2: f"{main_pset}:{main_attrib} fehlt",
                3: "Bauteilklassifikation nicht in SOM vorhanden",
                4: "GUID ist in mehreren Dateien identisch",
                5: "PropertySet Fehlt",
                6: "Attribut Fehlt",
                7: "Attribut hat falschen Wert",
                8: "Gruppe hat falsches Subelement",
                9: "Zwischenebene besitzt verschiedene Klassen als Subelement",
                10: "Gruppe besitzt keine Subelemente",
                11: "Element hat keine Gruppenzuweisen",
                12: "Zu Viele Subelemente"}

    def start_task(self) -> tuple:
        proj, ifc, pset, attribute, export_path = super(ModelcheckWidget, self).start_task()
        self.data_base_path = tempfile.NamedTemporaryFile().name
        logging.info(f"Database: {self.data_base_path}")
        window: ModelcheckWindow = self.window()
        data_dict = window.build_data_dict()

        self.runner = Modelcheck(ifc, proj, pset, attribute, export_path, self.data_base_path, data_dict)
        self.connect_runner(self.runner)
        self.thread_pool.start(self.runner)
        return proj, ifc, pset, attribute, export_path


class Modelcheck(IfcRunner):

    def __init__(self, ifc_paths: str, project: classes.Project, main_pset: str, main_attribute: str, issue_path: str,
                 data_base_path: str,
                 data_dict: dict[classes.Object, dict[classes.PropertySet, list[classes.Attribute]]]):
        super(Modelcheck, self).__init__(ifc_paths, project, main_pset, main_attribute, "Modelcheck")
        self.data_base_path = data_base_path
        self.base_name: str = ""
        self.ident_dict = dict()
        self.group_dict: dict[ifcopenshell.entity_instance] = dict()
        self.group_parent_dict = dict()
        self.export_path = issue_path
        self.data_dict = data_dict

    def run(self) -> None:
        """Iterates over all files and runs run_file_function for them"""
        sql.guids = dict()
        sql.create_tables(self.data_base_path)
        super(Modelcheck, self).run()
        if self.is_aborted:
            return
        self.create_issues()

    def run_file_function(self, file_path) -> ifcopenshell.file:

        ifc_file = super(Modelcheck, self).run_file_function(file_path)
        self.check_all_elements(ifc_file)
        return ifc_file

    def _get_ident_dict(self) -> dict[str:classes.Object]:
        return {obj.ident_value: obj for obj in self.project.objects}

    @staticmethod
    def _get_root_groups(ifc: ifcopenshell.file) -> list[ifcopenshell.entity_instance]:
        return [group for group in ifc.by_type("IfcGroup") if not modelcheck.get_parent_group(group)]

    def _build_group_structure(self, ifc: ifcopenshell.file):
        group_dict = dict()
        group_parent_dict = dict()
        for entity in self._get_root_groups(ifc):
            group_dict[entity] = dict()
            modelcheck.iterate_group_structure(entity, group_dict[entity], self.main_pset, self.main_attribute,
                                               group_parent_dict)
        return group_dict, group_parent_dict

    def _reset_checked_objects(self) -> None:
        self.checked_objects = 0

    def check_all_elements(self, ifc: ifcopenshell.file):
        self.signaller.status.emit("Prüfe Elemente mit Gruppenzuordnung")
        sql.remove_existing_issues(self.data_base_path, self.project.name, datetime.today(), self.base_name)
        self.ident_dict = self._get_ident_dict()
        self.group_dict, self.group_parent_dict = self._build_group_structure(ifc)
        self._reset_checked_objects()
        self.object_count = count_dict(self.group_dict)
        self.check_groups(self.group_dict)

        self.signaller.status.emit("Prüfe Elemente ohne Gruppenzuordnung")
        entites_without_group = _get_entities_without_group_assertion(ifc)
        self._reset_checked_objects()
        self.object_count = len(entites_without_group)
        self.check_elements_witout_group_assertion(entites_without_group)

    def create_issues(self):
        def get_max_width():
            col_widths = [0 for _ in range(worksheet.max_column)]
            for r_index, row in enumerate(worksheet, start=1):
                for c_index, cell in enumerate(row):
                    length = len(str(cell.value))
                    if length > col_widths[c_index]:
                        col_widths[c_index] = length
            return col_widths

        directory = os.path.dirname(self.export_path)
        if not os.path.exists(directory):
            os.mkdir(directory)

        _issues = sql.query_issues(self.data_base_path)

        self.signaller.status.emit(f"{len(_issues)} Fehler gefunden!")

        if len(_issues) == 0:
            self.signaller.status.emit("Modelle fehlerfrei!")
            return

        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        for col_index, value in enumerate(HEADER, start=1):
            worksheet.cell(1, col_index, value)

        last_cell = worksheet.cell(1, 8)

        for row_index, column in enumerate(_issues, start=2):
            for column_index, value in enumerate(column, start=1):
                last_cell = worksheet.cell(row_index, column_index, value)  # remove Whitespace

        table_zone = f"A1:{last_cell.coordinate}"
        tab = Table(displayName="Issues", ref=table_zone)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False,
                               showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        worksheet.add_table(tab)

        max_widths = get_max_width()

        # autoFit Column
        dim_holder = DimensionHolder(worksheet=worksheet)
        for col in range(worksheet.min_column, worksheet.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(worksheet, min=col, max=col,
                                                                 width=max_widths[col - 1] * 1.1)
        worksheet.column_dimensions = dim_holder

        self.save_workbook(workbook, self.export_path)

    def save_workbook(self, workbook, path):
        try:
            workbook.save(path)
        except PermissionError:
            logging.warning("-" * 60)
            logging.warning(f"folgende Datei ist noch geöffnet: {path} \n Datei schließen und beliebige Taste Drücken")
            input("Achtung! Datei wird danach überschrieben!")
            self.save_workbook(workbook, path)

    def check_groups(self, group_dict: dict):
        for entity in group_dict:
            if self.is_aborted:
                return
            self.check_single_group(entity, self.group_dict, 0)

    def check_single_group(self, element: ifcopenshell.entity_instance, group_dict, layer_index: int):

        def check_collector_group():
            sql.db_create_entity(self.data_base_path, element, self.project.name, self.base_name, identifier)
            for sub_ident in sub_idents:
                if sub_ident != identifier:
                    issues.subgroup_issue(self.data_base_path, element.GlobalId, sub_ident)

        def check_real_element():
            """checks group or element that is not a collector"""

            def loop_parent(el: classes.Aggregation) -> classes.Aggregation:
                if el.parent_connection != value_constants.INHERITANCE:
                    return el.parent
                else:
                    return loop_parent(el.parent)

            def check_correct_parent():
                parent_group = self.group_parent_dict.get(self.group_parent_dict.get(element))
                allowed_parents = set(loop_parent(aggreg) for aggreg in object_rep.aggregations)
                _allowed = [aggreg.name for aggreg in allowed_parents if aggreg is not None]
                logging.info(
                    f"Group {element.GlobalId} Allowed parents= {_allowed}")
                if parent_group is None:
                    if None not in allowed_parents:
                        logging.warning(f"Group {element.GlobalId} -> no parent group")
                    return
                parent_object = self.ident_dict.get(
                    modelcheck.get_identifier(parent_group, self.main_pset, self.main_attribute))
                if parent_object is None:
                    logging.warning(f"Group {element.GlobalId} -> no parent obj")
                    return

                parent_in_allowed_parents = bool(parent_object.aggregations.union().intersection(allowed_parents))
                if not (parent_in_allowed_parents or None in allowed_parents):
                    issues.parent_issue(self.data_base_path, element, parent_group, self.main_pset, self.main_attribute,
                                        element_type)

            def check_repetetive_group():
                """Gruppe besitzt mehrere Subelemente mit der selben Bauteilklassifikation"""
                if len(sub_idents) != len(
                        [modelcheck.get_identifier(sub_group, self.main_pset, self.main_attribute) for sub_group in
                         group_dict[element]]):
                    logging.warning(
                        f"Gruppe {element.GlobalId} -> "
                        f"besitzt mehrere Subelemente mit der selben Bauteilklassifikation")
                    issues.repetetive_group_issue(self.data_base_path, element)

            modelcheck.check_element(element, self.main_pset, self.main_attribute, self.data_base_path, self.base_name,
                                     self.ident_dict, element_type, self.project.name, self.data_dict)

            if object_rep is None:
                logging.warning(f"Group {element.GlobalId} -> no obj rep")
                return
            check_correct_parent()
            if element_type == modelcheck.GROUP:
                check_repetetive_group()

        if self.is_aborted:
            return

        self.increment_progress("Prüfe Elemente mit Gruppenzuordnung")

        identifier = modelcheck.get_identifier(element, self.main_pset, self.main_attribute)
        even_layer = layer_index % 2 == 0
        object_rep = self.ident_dict.get(identifier)
        if object_rep is not None and object_rep not in self.data_dict:
            return

        if element.is_a("IfcElement"):
            element_type = modelcheck.ELEMENT
            check_real_element()
            return
        else:
            element_type = modelcheck.GROUP

        sub_idents = {modelcheck.get_identifier(sub_group, self.main_pset, self.main_attribute) for sub_group in
                      group_dict[element]}
        sub_entities = group_dict[element].keys()

        if even_layer:
            check_collector_group()
        else:
            check_real_element()

        if len(sub_entities) == 0:
            issues.empty_group_issue(self.data_base_path, element)

        for sub_group in group_dict[element]:
            if self.is_aborted:
                return
            self.check_single_group(sub_group, group_dict[element], layer_index + 1)

    def check_elements_witout_group_assertion(self, entites_without_group: list[ifcopenshell.entity_instance]):
        for entity in entites_without_group:
            if self.is_aborted:
                return
            self.increment_progress("Prüfe Elemente ohne Gruppenzuordnung", 1)

            modelcheck.check_element(entity, self.main_pset, self.main_attribute, self.data_base_path, self.base_name,
                                     self.ident_dict, modelcheck.ELEMENT, self.project.name, self.data_dict, False)


def _get_entities_without_group_assertion(ifc: ifcopenshell.file) -> list[ifcopenshell.entity_instance]:
    return [entity for entity in ifc.by_type("IfcElement") if
            not [assignment for assignment in getattr(entity, "HasAssignments", []) if
                 assignment.is_a("IfcRelAssignsToGroup")]]
