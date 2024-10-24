from __future__ import annotations

import os
import sqlite3
from typing import TYPE_CHECKING

import openpyxl
from PySide6.QtCore import QCoreApplication
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

import som_gui
import som_gui.core.tool
from som_gui import tool
from som_gui.module.modelcheck_results import trigger

if TYPE_CHECKING:
    from som_gui.module.modelcheck_results.prop import ModelcheckResultProperties


class ModelcheckResults(som_gui.core.tool.ModelcheckResults):
    @classmethod
    def get_header(cls) -> list[str]:
        header = [
            QCoreApplication.translate("Modelcheck", "Datum"),
            QCoreApplication.translate("Modelcheck", "GUID"),
            QCoreApplication.translate("Modelcheck", "IfcType"),
            QCoreApplication.translate("Modelcheck", "Decription"),
            QCoreApplication.translate("Modelcheck", "Issue-Type"),
            QCoreApplication.translate("Modelcheck", "Name"),
            QCoreApplication.translate("Modelcheck", "Identifier"),
            QCoreApplication.translate("Modelcheck", "PropertySet"),
            QCoreApplication.translate("Modelcheck", "Attribute"),
            QCoreApplication.translate("Modelcheck", "Value"),
            QCoreApplication.translate("Modelcheck", "File"),

        ]
        return header

    @classmethod
    def last_modelcheck_finished(cls, ):
        trigger.last_modelcheck_finished(tool.Modelcheck.get_database_path())

    @classmethod
    def autofit_column_width(cls, worksheet: Worksheet):
        max_widths = cls.get_max_width(worksheet)
        dim_holder = DimensionHolder(worksheet=worksheet)
        for col in range(worksheet.min_column, worksheet.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(worksheet, min=col, max=col,
                                                                 width=max_widths[col - 1] * 1.1)
        worksheet.column_dimensions = dim_holder

    @classmethod
    def create_table(cls, worksheet: Worksheet, last_cell: Cell):
        table_zone = f"A1:{last_cell.coordinate}"
        tab = Table(displayName="Issues", ref=table_zone)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False,
                               showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        worksheet.add_table(tab)

    @classmethod
    def fill_worksheet(cls, issues, ws: Worksheet):
        last_cell = ws.cell(1, len(cls.get_header()) - 1)
        for row_index, column in enumerate(issues, start=2):
            for column_index, value in enumerate(column, start=1):
                last_cell = ws.cell(row_index, column_index, value)  # remove Whitespace
        return last_cell

    @classmethod
    def create_workbook(cls) -> tuple[openpyxl.Workbook, Worksheet]:

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        for col_index, value in enumerate(cls.get_header(), start=1):
            worksheet.cell(1, col_index, value)
        return workbook, worksheet

    @classmethod
    def get_properties(cls) -> ModelcheckResultProperties:
        return som_gui.ModelcheckResultProperties

    @classmethod
    def get_max_width(cls, worksheet: Worksheet) -> list[int]:
        col_widths = [0 for _ in range(worksheet.max_column)]
        for row in worksheet:
            for c_index, cell in enumerate(row):
                length = len(str(cell.value))
                if length > col_widths[c_index]:
                    col_widths[c_index] = length
        return col_widths

    @classmethod
    def get_export_path(cls):
        return cls.get_properties().excel_export_path

    @classmethod
    def set_export_path(cls, value: str):
        cls.get_properties().excel_export_path = value

    @classmethod
    def query_issues(cls, path: os.PathLike | str) -> list:
        conn = sqlite3.connect(path)
        cursor = conn.cursor()
        cursor.execute(
            'SELECT i.creation_date, e.GUID,e.ifc_type,i.short_description,i.issue_type,e.Name,e.bauteilKlassifikation,'
            'i.PropertySet,i.Attribut,i.Value, e.datei   FROM issues AS i JOIN entities e on i.GUID = e.GUID')
        conn.commit()
        query = cursor.fetchall()
        return query
